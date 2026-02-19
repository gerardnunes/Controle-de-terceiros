# --- Gerente: Aprovação de Chamadas ---
from datetime import timedelta, timezone
from itertools import count
from pyexpat.errors import messages
from dash import Output
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.contrib import messages
from django.db.models import Count
from django.db.models import Count, F, ExpressionWrapper, IntegerField, Sum
from django.db.models import DecimalField
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from django.db.models import Sum  
from gestao.decorators import role_required
from gestao.forms import RelatorioForm
from gestao.models import Chamada, User

import openpyxl
from io import BytesIO
from openpyxl import Workbook
from django.http import HttpResponse

from django.http import HttpResponse
from django.utils import timezone

@login_required
@role_required('gerente')
def exportar_quinzena_excel(request):
    # Dados da quinzena (mesmo filtro do dashboard)
    hoje = timezone.now().date()
    quinze_dias_atras = hoje - timedelta(days=15)
    chamadas = Chamada.objects.filter(
        data__gte=quinze_dias_atras,
        status='concluida'
    ).select_related('encarregado').order_by('-data')

    # Criar workbook e planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório Quinzenal"

    # Cabeçalhos
    cabecalhos = ['Data', 'Encarregado', 'Valor (R$)']
    ws.append(cabecalhos)

    # Dados
    total = 0
    for chamada in chamadas:
        nome = f"{chamada.encarregado.first_name} {chamada.encarregado.last_name}"
        ws.append([
            chamada.data.strftime("%d/%m/%Y"),
            nome,
            float(chamada.valor) if chamada.valor else 0
        ])
        total += chamada.valor or 0

    # Linha de total
    ws.append(['TOTAL', '', total])

    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=relatorio_quinzena.xlsx'

    return response

@login_required
@role_required('gerente')
def gerente_chamada_detail(request, pk):
    chamada = get_object_or_404(Chamada, pk=pk)
    if request.method == 'POST':
        acao = request.POST.get('acao')
        if acao == 'aprovar':
            chamada.status = 'aprovado'
            chamada.aprovado_por = request.user
            chamada.aprovado_em = timezone.now()
            chamada.save()
            messages.success(request, 'Chamada aprovada.')
        elif acao == 'rejeitar':
            chamada.status = 'rejeitado'
            chamada.aprovado_por = request.user
            chamada.aprovado_em = timezone.now()
            chamada.save()
            messages.success(request, 'Chamada rejeitada.')
        return redirect('dashboard')
    presencas = chamada.presencas.all()
    return render(request, 'gerente/chamada_detail.html', {'chamada': chamada, 'presencas': presencas})

# --- Relatórios (Gerente e Gestor) ---
@login_required
@role_required('gerente', 'gestor')
def relatorio(request):
    # Recebe as datas via GET
    inicio = request.GET.get('inicio')
    fim = request.GET.get('fim')
    if inicio and fim:
        # Converte string para date
        from datetime import datetime
        inicio = datetime.strptime(inicio, '%Y-%m-%d').date()
        fim = datetime.strptime(fim, '%Y-%m-%d').date()
    else:
        # Valores padrão (últimos 30 dias, por exemplo)
        from django.utils import timezone
        fim = timezone.now().date()
        inicio = fim - timedelta(days=30)

    # Filtrar chamadas concluídas no período

    chamadas = Chamada.objects.filter(
        data__range=[inicio, fim],# ajuste conforme seu modelo
    ).select_related('encarregado')
    VALOR_POR_CHAMADA = 120

    chamadas = Chamada.objects.filter(
        data__range=[inicio, fim],
        status='aprovado'  # recomendo filtrar só aprovadas
    ).prefetch_related('presencas__usuario')

    dados_agrupados = chamadas.values(
        'presencas__usuario__id',
        'presencas__usuario__first_name',
        'presencas__usuario__last_name',
    ).annotate(
        total_chamadas=Count('presencas'),
        total_valor=ExpressionWrapper(
            Count('presencas') * VALOR_POR_CHAMADA,
         output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    ).order_by('-total_valor')


    # Preparar lista para o template, calculando dias trabalhados e média
    dados_template = []
    total_dias = 0
    total_chamadas = 0
    total_valor = 0

    for item in dados_agrupados:
        # Dias trabalhados: contar dias distintos em que o usuário teve chamadas
        dias_trabalhados = chamadas.filter(
            presencas__usuario__id=item['presencas__usuario__id']
        ).values('data').distinct().count()


        nome_completo = f"{item['presencas__usuario__first_name']} {item['presencas__usuario__last_name']}"
        valor = item['total_valor'] or 0
        media = valor / dias_trabalhados if dias_trabalhados else 0

        dados_template.append({
            'first_name': nome_completo,
            'dias': dias_trabalhados,
            'total_chamadas': item['total_chamadas'],
            'valor': valor,
            'media_diaria': media,
        })

        total_dias += dias_trabalhados
        total_chamadas += item['total_chamadas']
        total_valor += valor

    totais = {
        'dias': total_dias,
        'chamadas': total_chamadas,
        'valor': total_valor,
        'media_geral': total_valor / total_dias if total_dias else 0,
    }

    context = {
        'inicio': inicio,
        'fim': fim,
        'dados': dados_template,
        'chamadas_periodo': chamadas.order_by('data'),  # para o detalhamento diário
        'totais': totais,
    }
    return render(request, 'gerente/relatorio_result.html', context)




#ESSE QUE TA EXPORTANDO ERRADO!!!
def exportar_relatorio_excel(request):
    from datetime import datetime
    from django.db.models import Count, DecimalField, ExpressionWrapper
    from openpyxl.utils import get_column_letter

    VALOR_POR_CHAMADA = 120

    # Datas
    inicio = request.GET.get('inicio')
    fim = request.GET.get('fim')

    if inicio and fim:
        inicio = datetime.strptime(inicio, '%Y-%m-%d').date()
        fim = datetime.strptime(fim, '%Y-%m-%d').date()
    else:
        fim = timezone.now().date()
        inicio = fim - timedelta(days=30)

    # Buscar chamadas
    chamadas = Chamada.objects.filter(
        data__range=[inicio, fim],
        status='aprovado'
    ).prefetch_related('presencas__usuario')

    # Agrupar por usuário
    dados_agrupados = chamadas.values(
        'presencas__usuario__id',
        'presencas__usuario__first_name',
        'presencas__usuario__last_name',
    ).annotate(
        total_chamadas=Count('presencas'),
        total_valor=ExpressionWrapper(
            Count('presencas') * VALOR_POR_CHAMADA,
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    ).order_by('-total_valor')

    wb = openpyxl.Workbook()

    # =========================
    # ABA 1 — RESUMO
    # =========================
    ws_resumo = wb.active
    ws_resumo.title = "Resumo por Usuário"

    ws_resumo.append([
        'Usuário',
        'Dias Trabalhados',
        'Total Chamadas',
        'Valor Total (R$)',
        'Média Diária (R$)'
    ])

    total_dias = 0
    total_chamadas = 0
    total_valor = 0

    for item in dados_agrupados:

        usuario_id = item['presencas__usuario__id']

        dias_trabalhados = chamadas.filter(
            presencas__usuario__id=usuario_id
        ).values('data').distinct().count()

        nome = f"{item['presencas__usuario__first_name']} {item['presencas__usuario__last_name']}"
        valor = item['total_valor'] or 0
        media = valor / dias_trabalhados if dias_trabalhados else 0

        ws_resumo.append([
            nome,
            dias_trabalhados,
            item['total_chamadas'],
            float(valor),
            round(media, 2)
        ])

        total_dias += dias_trabalhados
        total_chamadas += item['total_chamadas']
        total_valor += valor

    ws_resumo.append([
        'TOTAIS',
        total_dias,
        total_chamadas,
        float(total_valor),
        round(total_valor / total_dias, 2) if total_dias else 0
    ])

    for col in ws_resumo.columns:
        max_len = max(len(str(cell.value)) for cell in col) + 2
        ws_resumo.column_dimensions[get_column_letter(col[0].column)].width = max_len

    # =========================
    # ABA 2 — DETALHAMENTO
    # =========================
    ws_detalhe = wb.create_sheet("Detalhamento Diário")
    ws_detalhe.append(['Data', 'Usuário', 'Valor (R$)'])

    for chamada in chamadas.order_by('data'):
        for presenca in chamada.presencas.all():
            usuario = presenca.usuario
            nome = f"{usuario.first_name} {usuario.last_name}"

            ws_detalhe.append([
                chamada.data.strftime("%d/%m/%Y"),
                nome,
                VALOR_POR_CHAMADA
            ])

    for col in ws_detalhe.columns:
        max_len = max(len(str(cell.value)) for cell in col) + 2
        ws_detalhe.column_dimensions[get_column_letter(col[0].column)].width = max_len

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=relatorio.xlsx'

    return response
