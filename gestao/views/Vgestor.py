 #--- Gestor: Gerenciar todos os usuários ---
from pyexpat.errors import messages
from io import BytesIO
from urllib import response
from urllib import response

output = BytesIO()
from django.http import HttpResponse
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
import openpyxl
from gestao.decorators import role_required
from gestao.forms import UsuarioForm
from gestao.models import Chamada, User
from datetime import datetime, timedelta
from django.db.models import Count, DecimalField, ExpressionWrapper
from openpyxl.utils import get_column_letter
from django.utils import timezone

@login_required
@role_required('gestor')
def gestor_usuario_list(request):
    usuarios = User.objects.all().order_by('role', 'username')
    return render(request, 'gestor/usuario_list.html', {'usuarios': usuarios})

@login_required
@role_required('gestor')
def gestor_usuario_edit(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuário atualizado.')
            return redirect('gestor_usuario_list')
    else:
        form = UsuarioForm(instance=user)
    return render(request, 'gestor/usuario_form.html', {'form': form})

@login_required
@role_required('gestor')
def gestor_usuario_list(request):
    usuarios_total = User.objects.filter(role='usuario').count()
    usuarios = User.objects.all().order_by('role', 'username')
    total_usuarios = User.objects.count()

    return render(request, 'gestor/dashboard.html', {'usuarios': usuarios, 'usuarios_total': usuarios_total, 'total_usuarios': total_usuarios})

@login_required
@role_required('gestor')
def usuario_p_setor(request):
    inicio = request.GET.get('inicio')
    fim = request.GET.get('fim')

    if inicio and fim:
        inicio = datetime.strptime(inicio, '%Y-%m-%d').date()
        fim = datetime.strptime(fim, '%Y-%m-%d').date()
    else:
        fim = timezone.now().date()
        inicio = fim - timedelta(days=30)

    # Buscar chamadas
    chamadas = Chamada.objects.filter(
        data__range=[inicio, fim],
        status='aprovado'
    ).prefetch_related('presencas__usuario')
    VALOR_POR_CHAMADA = 120
    # Agrupar por usuário
    dados_agrupados = chamadas.values(
        'presencas__usuario__id',
        'presencas__usuario__first_name',
        'presencas__local__nome',
    )

    wb = openpyxl.Workbook()

    # =========================
    # ABA 1 — RESUMO
    # =========================
    ws_resumo = wb.active
    ws_resumo.title = "Resumo por Usuário"

    ws_resumo.append([
        'Usuário',
        'Dias Trabalhados',
        'Total Chamadas',
            ])

    total_dias = 0
    total_chamadas = 0
    total_valor = 0

    for item in dados_agrupados:
        usuario_id = item['presencas__usuario__id']
        local = item['presencas__local__nome']
        dias_trabalhados = chamadas.filter(
            presencas__usuario__id=usuario_id
        ).values('data').distinct().count()

        nome = f"{item['presencas__usuario__first_name']}"

        ws_resumo.append([
            nome,
            dias_trabalhados,
            local,
        ])

        total_dias += dias_trabalhados


    for col in ws_resumo.columns:
        max_len = max(len(str(cell.value)) for cell in col) + 2
        ws_resumo.column_dimensions[get_column_letter(col[0].column)].width = max_len

    # =========================
    # ABA 2 — DETALHAMENTO
    # =========================

    # =========================
    from io import BytesIO

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=relatorio.xlsx'

    return response


@login_required
@role_required('gestor')
def chamada_periodo(request):
    inicio = request.GET.get('inicio')
    fim = request.GET.get('fim')

    if inicio and fim:
        inicio = datetime.strptime(inicio, '%Y-%m-%d').date()
        fim = datetime.strptime(fim, '%Y-%m-%d').date()
    else:
        fim = timezone.now().date()
        inicio = fim - timedelta(days=30)

    # Buscar chamadas
    #chamadas_hoje = Chamada.objects.filter(data=hoje).first()
    chamadas = Chamada.objects.filter(
        data__range=[inicio, fim],
        status='aprovado'
    )
    VALOR_POR_CHAMADA = 120
    # Agrupar por usuário
    dados_agrupados = chamadas.values(
        'aprovado_em',
        'data',
        'presencas',
        'presencas__local__nome',
    )

    wb = openpyxl.Workbook()

    ws_resumo = wb.active
    ws_resumo.title = "Resumo por Usuário"

    ws_resumo.append([
        'Usuário',
        'Dias Trabalhados',
        'Total Chamadas',
        "testes"
            ])

    total_dias = 0

    from django.utils.timezone import localtime
    for item in dados_agrupados:
        aprovado_em = item['aprovado_em']
        data = item['data']

        if isinstance(aprovado_em, datetime):
            aprovado_em = localtime(aprovado_em).replace(tzinfo=None)


        nome = item['presencas__local__nome']

        ws_resumo.append([
            nome,
            aprovado_em,
            data,
        ])

    for col in ws_resumo.columns:
        max_len = max(len(str(cell.value)) for cell in col) + 2
        ws_resumo.column_dimensions[get_column_letter(col[0].column)].width = max_len
    from io import BytesIO

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=relatorio.xlsx'

    return response